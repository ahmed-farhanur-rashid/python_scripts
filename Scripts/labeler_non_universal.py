# Dataset Labeler
# A simple utility to generate CSV and XLSX files containing image paths and labels based on folder structure down below.
#
#   Dataset
#   ├─A
#   │ ├─a
#   │ │ ├─i
#   │ │ └─ii
#   │ └─b
#   │    ├─i
#   │    └─ii
#   └─B
#      ├─a
#      │ ├─i
#      │ └─ii
#      └─b
#         ├─i
#         └─ii
# The above structure with HEADERS = ["image_name", "relative_path", "license_type", "crop_type", "quality"]

import os
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ══════════════════════════════════════════════════════════════════════════════
#  CONFIGURATION  ─ only section you ever need to touch
# ══════════════════════════════════════════════════════════════════════════════

# First two headers are always image_name and relative_path.
# The remaining headers map 1-to-1 to each subfolder depth level (level 1, 2, 3 …).
HEADERS = [
    "image_name",
    "relative_path",
    "license_type",   # depth 1 folder name
    "crop_type",      # depth 2 folder name
    "quality",        # depth 3 folder name
]

OUTPUT_STEM = "dataset"   # produces dataset.csv and dataset.xlsx

# ══════════════════════════════════════════════════════════════════════════════
#  CORE LOGIC
# ══════════════════════════════════════════════════════════════════════════════

# The number of folder-depth columns is everything after image_name + relative_path
DEPTH = len(HEADERS) - 2


def collect_rows(base_dir: str) -> list[dict]:
    """
    Recursively walk `base_dir` to exactly `DEPTH` subfolder levels, then
    collect every file found at that leaf level.  Folder names at each level
    become the label values — no whitelist required.
    """
    rows: list[dict] = []

    def _walk(current_dir: str, depth: int, labels: list[str]) -> None:
        entries = sorted(os.listdir(current_dir))
        if depth < DEPTH:
            for entry in entries:
                child = os.path.join(current_dir, entry)
                if os.path.isdir(child):
                    _walk(child, depth + 1, labels + [entry])
        else:
            for fname in entries:
                fpath = os.path.join(current_dir, fname)
                if not os.path.isfile(fpath):
                    continue
                rel = os.path.join(*labels, fname)
                row = {"image_name": fname, "relative_path": rel}
                for header, label in zip(HEADERS[2:], labels):
                    row[header] = label
                rows.append(row)

    _walk(base_dir, 0, [])
    return rows


# ══════════════════════════════════════════════════════════════════════════════
#  CSV WRITER
# ══════════════════════════════════════════════════════════════════════════════

def write_csv(rows: list[dict], path: str) -> None:
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=HEADERS)
        writer.writeheader()
        writer.writerows(rows)
    print(f"  CSV  -> {path}  ({len(rows)} rows)")


# ══════════════════════════════════════════════════════════════════════════════
#  XLSX WRITER
# ══════════════════════════════════════════════════════════════════════════════

_HEADER_FILL  = PatternFill("solid", start_color="2F5496")
_HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_CELL_FONT    = Font(name="Arial", size=10)
_ALIGN_L      = Alignment(horizontal="left",   vertical="center")
_ALIGN_C      = Alignment(horizontal="center", vertical="center")
_ODD_FILL     = PatternFill("solid", start_color="EEF2F9")
_EVEN_FILL    = PatternFill("solid", start_color="FFFFFF")
_THIN         = Side(style="thin", color="BFBFBF")
_BORDER       = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_CENTER_FROM  = 3   # 1-indexed; image_name & relative_path stay left-aligned


def _col_width(header: str, rows: list[dict]) -> float:
    max_len = len(header)
    for row in rows[:500]:
        val = row.get(header, "")
        if val:
            max_len = max(max_len, len(str(val)))
    return min(max_len + 4, 60)


def _build_summary(rows: list[dict]) -> list[tuple]:
    summary: list[tuple] = [("Total images", len(rows))]
    for header in HEADERS[2:]:
        counts: dict[str, int] = {}
        for row in rows:
            val = row.get(header, "")
            counts[val] = counts.get(val, 0) + 1
        for val, count in sorted(counts.items()):
            summary.append((f"{header} = {val}", count))
    return summary


def write_xlsx(rows: list[dict], path: str) -> None:
    wb = openpyxl.Workbook()

    # Dataset sheet
    ws = wb.active
    ws.title = "Dataset"
    ws.row_dimensions[1].height = 28

    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill      = _HEADER_FILL
        cell.font      = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN
        cell.border    = _BORDER
        ws.column_dimensions[cell.column_letter].width = _col_width(header, rows)

    for row_idx, record in enumerate(rows, start=2):
        fill = _ODD_FILL if row_idx % 2 == 1 else _EVEN_FILL
        for col_idx, header in enumerate(HEADERS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=record.get(header, ""))
            cell.font      = _CELL_FONT
            cell.fill      = fill
            cell.border    = _BORDER
            cell.alignment = _ALIGN_C if col_idx >= _CENTER_FROM else _ALIGN_L

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(HEADERS))}1"

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 14
    for col, val in [(1, "Breakdown"), (2, "Count")]:
        c = ws2.cell(row=1, column=col, value=val)
        c.font = Font(name="Arial", bold=True, size=11)

    for r_idx, (label, value) in enumerate(_build_summary(rows), start=2):
        lc = ws2.cell(row=r_idx, column=1, value=label)
        vc = ws2.cell(row=r_idx, column=2, value=value)
        for cell in (lc, vc):
            cell.font   = Font(name="Arial", size=10)
            cell.border = _BORDER
        lc.alignment = _ALIGN_L
        vc.alignment = _ALIGN_C

    wb.save(path)
    print(f"  XLSX -> {path}  ({len(rows)} rows, {len(HEADERS)} columns)")


# ══════════════════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    base = os.path.dirname(os.path.abspath(__file__))

    print(f"Scanning : {base}")
    print(f"Depth    : {DEPTH} level(s)")
    print(f"Labels   : {HEADERS[2:]}\n")

    rows = collect_rows(base)

    if not rows:
        print("No files found. Make sure the folder depth matches the number of label headers.")
    else:
        print(f"Found {len(rows)} file(s). Writing outputs ...\n")
        write_csv(rows,  os.path.join(base, f"{OUTPUT_STEM}.csv"))
        write_xlsx(rows, os.path.join(base, f"{OUTPUT_STEM}.xlsx"))
        print("\nDone.")
