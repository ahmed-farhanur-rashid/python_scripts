"""
generate_dataset.py
────────────────────────────────────────────────────────────────────────────────
Drop this script into any folder that contains a nested image dataset.

FIRST RUN  →  scans the folder structure, writes `dataset_config.json`, exits.
              Open that file, rename any column header you like, set enabled:false
              to drop a column entirely, then re-run.

SECOND RUN →  reads `dataset_config.json`, builds `dataset.csv` + `dataset.xlsx`.

Internal slot names (the keys in "columns"):
  filename   →  the image file name
  path       →  relative path from the dataset root to the file
  level_1    →  name of the folder at depth 1
  level_2    →  name of the folder at depth 2
  … and so on for however many levels exist

You can rename the "header" of any slot to anything you like.
You can set "enabled": false on any slot to exclude it from the output.
────────────────────────────────────────────────────────────────────────────────
"""

import os
import csv
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
CONFIG_PATH = os.path.join(BASE_DIR, "dataset_config.json")

IMAGE_EXTENSIONS = {
    ".jpg", ".jpeg", ".png", ".bmp", ".tiff", ".tif",
    ".webp", ".gif", ".heic", ".heif", ".svg",
}

# ══════════════════════════════════════════════════════════════════════════════
#  STRUCTURE DISCOVERY
# ══════════════════════════════════════════════════════════════════════════════

def _is_image(fname: str) -> bool:
    return os.path.splitext(fname)[1].lower() in IMAGE_EXTENSIONS


def discover_structure(base_dir: str) -> dict:
    """
    Walk the tree, find the deepest level that contains images,
    and return a config skeleton with all slots enabled and named by their slot key.
    """
    max_depth = 0
    sample_tree: dict = {}

    def _walk(current: str, depth: int, node: dict) -> None:
        nonlocal max_depth
        entries = sorted(os.listdir(current))
        images  = [e for e in entries if os.path.isfile(os.path.join(current, e)) and _is_image(e)]
        subdirs = [e for e in entries if os.path.isdir(os.path.join(current, e)) and not e.startswith(".")]

        if images:
            node["__images__"] = images[:3]
            max_depth = max(max_depth, depth)

        for d in subdirs:
            node[d] = {}
            _walk(os.path.join(current, d), depth + 1, node[d])

    _walk(base_dir, 0, sample_tree)

    # Build the columns dict: fixed slots first, then one per level
    columns = {
        "filename": {"header": "filename", "enabled": True},
        "path":     {"header": "path",     "enabled": True},
    }
    for i in range(max_depth):
        key = f"level_{i + 1}"
        columns[key] = {"header": key, "enabled": True}

    return {
        "max_depth":   max_depth,
        "columns":     columns,
        "output_stem": "dataset",
        "_sample_tree": sample_tree,
    }


# ══════════════════════════════════════════════════════════════════════════════
#  ROW COLLECTION
# ══════════════════════════════════════════════════════════════════════════════

def collect_rows(base_dir: str, max_depth: int, columns: dict) -> list[dict]:
    rows: list[dict] = []

    def _walk(current: str, depth: int, labels: list[str]) -> None:
        entries = sorted(os.listdir(current))
        images  = [e for e in entries if os.path.isfile(os.path.join(current, e)) and _is_image(e)]
        subdirs = [e for e in entries if os.path.isdir(os.path.join(current, e)) and not e.startswith(".")]

        for fname in images:
            padded = labels + [""] * (max_depth - len(labels))
            row: dict = {}

            # filename slot
            if columns.get("filename", {}).get("enabled", True):
                row[columns["filename"]["header"]] = fname

            # path slot
            if columns.get("path", {}).get("enabled", True):
                rel = os.path.join(*labels, fname) if labels else fname
                row[columns["path"]["header"]] = rel

            # level slots
            for i, label_val in enumerate(padded):
                key = f"level_{i + 1}"
                if columns.get(key, {}).get("enabled", True):
                    row[columns[key]["header"]] = label_val

            rows.append(row)

        for d in subdirs:
            _walk(os.path.join(current, d), depth + 1, labels + [d])

    _walk(base_dir, 0, [])
    return rows


def active_headers(columns: dict, max_depth: int) -> list[str]:
    """Return headers in slot order, skipping disabled ones."""
    ordered_keys = ["filename", "path"] + [f"level_{i+1}" for i in range(max_depth)]
    return [
        columns[k]["header"]
        for k in ordered_keys
        if k in columns and columns[k].get("enabled", True)
    ]


def label_headers(columns: dict, max_depth: int) -> list[str]:
    """Return only the level headers (for summary grouping), skipping disabled."""
    return [
        columns[f"level_{i+1}"]["header"]
        for i in range(max_depth)
        if columns.get(f"level_{i+1}", {}).get("enabled", True)
    ]


# ══════════════════════════════════════════════════════════════════════════════
#  CSV WRITER
# ══════════════════════════════════════════════════════════════════════════════

def write_csv(rows: list[dict], headers: list[str], path: str) -> None:
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)
    print(f"  CSV  -> {path}  ({len(rows)} rows)")


# ══════════════════════════════════════════════════════════════════════════════
#  XLSX WRITER
# ══════════════════════════════════════════════════════════════════════════════

_HEADER_FILL  = PatternFill("solid", start_color="2F5496")
_HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_CELL_FONT    = Font(name="Arial", size=10)
_ALIGN_L      = Alignment(horizontal="left",   vertical="center")
_ALIGN_C      = Alignment(horizontal="center", vertical="center")
_ODD_FILL     = PatternFill("solid", start_color="EEF2F9")
_EVEN_FILL    = PatternFill("solid", start_color="FFFFFF")
_THIN         = Side(style="thin", color="BFBFBF")
_BORDER       = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _col_width(header: str, rows: list[dict]) -> float:
    max_len = len(header)
    for row in rows[:500]:
        val = row.get(header, "")
        if val:
            max_len = max(max_len, len(str(val)))
    return min(max_len + 4, 60)


def _build_summary(rows: list[dict], lbl_headers: list[str]) -> list[tuple]:
    summary: list[tuple] = [("Total images", len(rows))]
    for header in lbl_headers:
        counts: dict[str, int] = {}
        for row in rows:
            val = row.get(header) or "(none)"
            counts[val] = counts.get(val, 0) + 1
        for val, count in sorted(counts.items()):
            summary.append((f"{header}  =  {val}", count))
    return summary


def write_xlsx(rows, all_headers, lbl_headers, path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dataset"
    ws.row_dimensions[1].height = 28

    for col_idx, header in enumerate(all_headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = _HEADER_FILL; cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN; cell.border = _BORDER
        ws.column_dimensions[cell.column_letter].width = _col_width(header, rows)

    # path slot is left-aligned; everything else centre-aligned
    path_header = all_headers[1] if len(all_headers) > 1 else None

    for row_idx, record in enumerate(rows, start=2):
        fill = _ODD_FILL if row_idx % 2 == 1 else _EVEN_FILL
        for col_idx, header in enumerate(all_headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=record.get(header, ""))
            cell.font = _CELL_FONT; cell.fill = fill; cell.border = _BORDER
            cell.alignment = _ALIGN_L if header == path_header else _ALIGN_C

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(all_headers))}1"

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["A"].width = 35
    ws2.column_dimensions["B"].width = 14
    for col, val in [(1, "Breakdown"), (2, "Count")]:
        c = ws2.cell(row=1, column=col, value=val)
        c.font = Font(name="Arial", bold=True, size=11)

    r_idx = 2
    prev_group = None
    for label, value in _build_summary(rows, lbl_headers):
        group = label.split("  =  ")[0] if "  =  " in label else None
        if group and group != prev_group and prev_group is not None:
            r_idx += 1   # blank spacer row
        prev_group = group
        lc = ws2.cell(row=r_idx, column=1, value=label)
        vc = ws2.cell(row=r_idx, column=2, value=value)
        is_total = label == "Total images"
        for cell in (lc, vc):
            cell.font   = Font(name="Arial", size=10, bold=is_total)
            cell.border = _BORDER
        lc.alignment = _ALIGN_L; vc.alignment = _ALIGN_C
        r_idx += 1

    wb.save(path)
    print(f"  XLSX -> {path}  ({len(rows)} rows, {len(all_headers)} columns)")


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════════════

def _print_tree(node: dict, indent: int = 0) -> None:
    for key, val in node.items():
        if key == "__images__":
            suffix = " …" if len(val) >= 3 else ""
            print(" " * indent + f"[images: {', '.join(val)}{suffix}]")
        else:
            print(" " * indent + f"📁 {key}/")
            _print_tree(val, indent + 4)


def main() -> None:
    # ── First run ────────────────────────────────────────────────────────────
    if not os.path.exists(CONFIG_PATH):
        print("No dataset_config.json found. Scanning folder structure …\n")
        config = discover_structure(BASE_DIR)

        if config["max_depth"] == 0:
            print("No images found in any subfolder.")
            return

        write_config = lambda c, p: open(p, "w").write(json.dumps(c, indent=2))
        write_config(config, CONFIG_PATH)

        print(f"Found {config['max_depth']} folder level(s).")
        print(f"\nConfig written to:\n  {CONFIG_PATH}\n")
        print("Edit the config to rename any 'header' value or set 'enabled': false")
        print("to drop a column, then re-run.\n")
        print("All available slots:\n")
        for key, val in config["columns"].items():
            print(f"  {key:12s}  →  header: \"{val['header']}\",  enabled: {val['enabled']}")
        print("\nDiscovered structure (sample):")
        _print_tree(config["_sample_tree"])
        return

    # ── Subsequent runs ──────────────────────────────────────────────────────
    with open(CONFIG_PATH, encoding="utf-8") as f:
        config = json.load(f)

    max_depth   = config["max_depth"]
    columns     = config["columns"]
    output_stem = config.get("output_stem", "dataset")

    all_hdrs = active_headers(columns, max_depth)
    lbl_hdrs = label_headers(columns, max_depth)

    print("Config loaded.")
    print(f"  Depth   : {max_depth} level(s)")
    print(f"  Columns : {all_hdrs}\n")

    rows = collect_rows(BASE_DIR, max_depth, columns)

    if not rows:
        print("No images found.")
        return

    print(f"Found {len(rows)} image(s). Writing outputs …\n")
    write_csv(rows,  all_hdrs, os.path.join(BASE_DIR, f"{output_stem}.csv"))
    write_xlsx(rows, all_hdrs, lbl_hdrs, os.path.join(BASE_DIR, f"{output_stem}.xlsx"))
    print("\nDone.")


if __name__ == "__main__":
    main()
